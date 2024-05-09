from openpyxl import load_workbook
from openpyxl.styles import NamedStyle

def copy_data_with_source_formatting(source_file, destination_file, source_sheet_name, destination_sheet_name, start_cell):
    # Load source workbook and sheet
    source_wb = load_workbook(source_file, data_only=True)
    source_sheet = source_wb[source_sheet_name]

    # Load destination workbook and sheet
    destination_wb = load_workbook(destination_file)
    destination_sheet = destination_wb[destination_sheet_name]

    # Copy data and formats from source to destination
    for row in source_sheet.iter_rows(min_row=start_cell[0], values_only=True):
        dest_row = []
        for cell in row:
            # Copy cell value
            dest_row.append(cell)
        destination_sheet.append(dest_row)

    # Copy formatting
    for row in destination_sheet.iter_rows(min_row=start_cell[0], max_row=destination_sheet.max_row):
        for dest_cell in row:
            src_cell = source_sheet.cell(row=dest_cell.row, column=dest_cell.column)
            dest_cell.font = copy_font(src_cell.font)
            dest_cell.alignment = copy_alignment(src_cell.alignment)
            dest_cell.border = copy_border(src_cell.border)
            dest_cell.fill = copy_fill(src_cell.fill)
            dest_cell.number_format = src_cell.number_format
            dest_cell.protection = src_cell.protection

    # Save destination workbook
    destination_wb.save(destination_file)

# Functions to copy formatting properties
def copy_font(font):
    new_font = NamedStyle(name='font_copy')
    for key, value in font.__dict__.items():
        if key[0] != "_":
            setattr(new_font, key, value)
    return new_font

def copy_alignment(alignment):
    new_alignment = NamedStyle(name='alignment_copy')
    for key, value in alignment.__dict__.items():
        if key[0] != "_":
            setattr(new_alignment, key, value)
    return new_alignment

def copy_border(border):
    new_border = NamedStyle(name='border_copy')
    for key, value in border.__dict__.items():
        if key[0] != "_":
            setattr(new_border, key, value)
    return new_border

def copy_fill(fill):
    new_fill = NamedStyle(name='fill_copy')
    for key, value in fill.__dict__.items():
        if key[0] != "_":
            setattr(new_fill, key, value)
    return new_fill

    # Save destination workbook

copy_data_with_source_formatting('blank_temp.xlsx', 'example.xlsx', 'PERSONAL MONTHLY BUDGET', 'Sheet2', (1, 1))