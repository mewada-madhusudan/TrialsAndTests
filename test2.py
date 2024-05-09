from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def find_table_ranges(file_path):
    wb = load_workbook(file_path, read_only=True)
    table_ranges = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        # Initialize variables to track current table range
        table_start_row = None
        table_start_col = None
        table_end_row = 0
        table_end_col = 0

        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                if cell_value is not None:
                    # If current cell is not empty, update table boundaries
                    if table_start_row is None:
                        table_start_row = row
                        table_start_col = col
                    else:
                        table_end_col = max(table_end_col, col)
                        table_end_row = max(table_end_row, row)
                elif table_start_row is not None:
                    # If current cell is empty and table has started, finalize table range
                    table_ranges.append((sheet_name,
                                         f"{get_column_letter(table_start_col)}{table_start_row}:{get_column_letter(table_end_col)}{table_end_row}"))
                    # Reset table boundaries
                    table_start_row = None
                    table_start_col = None
                    table_end_row = 0
                    table_end_col = 0

        # If table extends to the last row, finalize table range
        if table_start_row is not None:
            table_ranges.append((sheet_name,
                                 f"{get_column_letter(table_start_col)}{table_start_row}:{get_column_letter(table_end_col)}{max_row}"))

    return table_ranges


# Example usage
file_path = "example.xlsx"  # Change this to your Excel file path
ranges = find_table_ranges(file_path)
for sheet_name, table_range in ranges:
    print(f"Table range in '{sheet_name}': {table_range}")
