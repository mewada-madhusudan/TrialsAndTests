from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_data_ranges(file_path):
    wb = load_workbook(file_path, read_only=True)
    data_ranges = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column
        print(max_row,max_col)


        if max_row > 0 and max_col > 0:
            # Initialize variables to track current data range
            current_range_start = None
            current_range_end = None

            for row in sheet.iter_rows(min_row=1, min_col=1, max_row=max_row, max_col=max_col):
                for cell in row:
                    if cell.value:
                        # If a cell has a value, update current range boundaries
                        if current_range_start is None:
                            current_range_start = cell.coordinate
                        current_range_end = cell.coordinate
                    print(cell.value)


            # If a non-empty range was detected, add it to the list
            if current_range_start is not None and current_range_end is not None:
                data_ranges.append((sheet_name, f"{current_range_start}:{current_range_end}"))

    return data_ranges

# Example usage
file_path = "example.xlsx"  # Change this to your Excel file path
ranges = find_data_ranges(file_path)
for sheet_name, data_range in ranges:
    print(f"Data range in '{sheet_name}': {data_range}")
