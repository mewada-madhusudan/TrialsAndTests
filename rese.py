import time

import openpyxl


def find_formula_cells(file_path):
    # Load the Excel workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)

    # Iterate over each sheet in the workbook
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]

        # Iterate over each row in the sheet
        for row in sheet.iter_rows():
            # Iterate over each cell in the row
            for cell in row:
                # Check if the cell contains a formula
                if cell.data_type == 'f':
                    # Print the sheet name and coordinates of the cell with a formula
                    print(f"Sheet: {sheet_name}, Cell with Formula: {cell.coordinate}")


# Specify the path to your Excel file
excel_file_path = "monthly report.xlsx"
start_time = time.time()
# Call the function to find cells with formulas
find_formula_cells(excel_file_path)
end_time = time.time()

# Calculate the elapsed time
elapsed_time = end_time - start_time

print(f"Elapsed time: {elapsed_time} seconds")