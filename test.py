from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def find_table_ranges(file_path):
    wb = load_workbook(file_path, read_only=True)
    table_ranges = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        if max_row > 0 and max_col > 0:
            in_table = False
            table_start = None

            for col in range(1, max_col + 1):
                col_empty = all(sheet.cell(row=row, column=col).value is None for row in range(1, max_row + 1))

                if not in_table and not col_empty:
                    in_table = True
                    table_start = get_column_letter(col) + '1'
                elif in_table and col_empty:
                    in_table = False
                    table_ranges.append((sheet_name, f"{table_start}:{get_column_letter(col - 1)}{max_row}"))

            # If the last column contains data, close the table
            if in_table:
                table_ranges.append((sheet_name, f"{table_start}:{get_column_letter(max_col)}{max_row}"))

    return table_ranges

# Example usage
file_path = "example.xlsx"  # Change this to your Excel file path
ranges = find_table_ranges(file_path)
for sheet_name, table_range in ranges:
    print(f"Table range in '{sheet_name}': {table_range}")
