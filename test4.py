from openpyxl import load_workbook

def find_sub_tables(file_path):
    wb = load_workbook(file_path, read_only=True)
    sub_tables = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        max_row = sheet.max_row
        max_col = sheet.max_column

        in_sub_table = False
        sub_table_start_row = None
        sub_table_start_col = None

        for row in range(1, max_row + 1):
            row_empty = all(sheet.cell(row=row, column=col).value is None for col in range(1, max_col + 1))
            if not row_empty:
                # Non-empty row
                if not in_sub_table:
                    # Start of a new sub-table
                    in_sub_table = True
                    sub_table_start_row = row
                    sub_table_start_col = 1
            elif in_sub_table:
                # Empty row, end of sub-table
                in_sub_table = False
                sub_table_end_row = row - 1
                sub_table_end_col = max_col
                sub_tables.append((sheet_name, sub_table_start_row, 1, sub_table_end_row, max_col))

        # Check if the last row contains data, indicating the end of a sub-table
        if in_sub_table:
            sub_table_end_row = max_row
            sub_table_end_col = max_col
            sub_tables.append((sheet_name, sub_table_start_row, 1, sub_table_end_row, max_col))

        # Reset variables for column-wise check
        in_sub_table = False
        sub_table_start_row = None
        sub_table_start_col = None

        for col in range(1, max_col + 1):
            col_empty = all(sheet.cell(row=row, column=col).value is None for row in range(1, max_row + 1))
            if not col_empty:
                # Non-empty column
                if not in_sub_table:
                    # Start of a new sub-table
                    in_sub_table = True
                    sub_table_start_row = 1
                    sub_table_start_col = col
            elif in_sub_table:
                # Empty column, end of sub-table
                in_sub_table = False
                sub_table_end_col = col - 1
                sub_tables.append((sheet_name, 1, sub_table_start_col, max_row, sub_table_end_col))

        # Check if the last column contains data, indicating the end of a sub-table
        if in_sub_table:
            sub_table_end_col = max_col
            sub_tables.append((sheet_name, 1, sub_table_start_col, max_row, sub_table_end_col))

    return sub_tables

# Example usage
file_path = "example.xlsx"  # Change this to your Excel file path
sub_tables = find_sub_tables(file_path)
wb = load_workbook(file_path, read_only=True)
sheet = wb['Sheet1']
for sheet_name, start_row, start_col, end_row, end_col in sub_tables:
    print(f"Sub-table in '{sheet_name}': From ({start_row}, {start_col}) to ({end_row}, {end_col})")
    # print(sheet.cell(row=start_row, column=start_col).coordinate+":"+sheet.cell(row=end_row, column=end_col).coordinate)




