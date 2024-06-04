import openpyxl

# Load the Excel file
workbook = openpyxl.load_workbook('your_excel_file.xlsx')

# Access the specific worksheet
worksheet = workbook['Sheet1']  # Replace 'Sheet1' with your sheet name

# Get the maximum row number and column letters
max_row = worksheet.max_row
max_col = worksheet.max_column

# Initialize variables to track ranges
similar_formula_range = []
different_formula_ranges = []

# Initialize variables to track current formula and range
current_formula = None
current_range_start = None
current_range_end = None

# Iterate over each column
for col_num in range(1, max_col + 1):
    column_letter = openpyxl.utils.get_column_letter(col_num)
    # Iterate over each row in the column
    for row_num in range(1, max_row + 1):
        cell = worksheet[column_letter + str(row_num)]
        formula = cell.formula
        if formula == current_formula:
            # If the formula is the same as the current formula, extend the range
            current_range_end = (column_letter, row_num)
        else:
            # If the formula is different, add the current range to the appropriate list
            if current_formula:
                if len(similar_formula_range) > 0:
                    similar_formula_range.append((current_range_start, current_range_end))
                else:
                    different_formula_ranges.append((current_range_start, current_range_end))
            # Start a new range for the current formula
            current_formula = formula
            current_range_start = (column_letter, row_num)
            current_range_end = (column_letter, row_num)

# Add the last range to the appropriate list
if len(similar_formula_range) > 0:
    similar_formula_range.append((current_range_start, current_range_end))
else:
    different_formula_ranges.append((current_range_start, current_range_end))

# Print the ranges
print("Ranges of cells with similar formulas:")
for start, end in similar_formula_range:
    print(f"{start[0]}{start[1]}:{end[0]}{end[1]}")

print("\nRanges of cells with different formulas:")
for start, end in different_formula_ranges:
    print(f"{start[0]}{start[1]}:{end[0]}{end[1]}")

# Close the workbook
workbook.close()
