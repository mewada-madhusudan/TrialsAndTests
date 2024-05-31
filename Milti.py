import pandas as pd
import openpyxl
from multiprocessing import Pool, cpu_count

def read_excel_chunk(file_path, min_row, max_row):
    workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    sheet = workbook.active

    data = []
    formulas = []

    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=False):
        data_row = []
        formula_row = []
        for cell in row:
            data_row.append(cell.value)
            formula_row.append(cell.formula if cell.formula else cell.value)
        data.append(data_row)
        formulas.append(formula_row)

    return data, formulas

def extract_formulas_multiprocessing(file_path, num_processes=None):
    if num_processes is None:
        num_processes = cpu_count()

    workbook = openpyxl.load_workbook(file_path, data_only=False, read_only=True)
    sheet = workbook.active
    total_rows = sheet.max_row

    chunk_size = total_rows // num_processes
    tasks = []

    for i in range(num_processes):
        min_row = i * chunk_size + 1
        max_row = (i + 1) * chunk_size if i != num_processes - 1 else total_rows
        tasks.append((file_path, min_row, max_row))

    with Pool(processes=num_processes) as pool:
        results = pool.starmap(read_excel_chunk, tasks)

    # Combine results
    data_combined = []
    formulas_combined = []

    for data, formulas in results:
        data_combined.extend(data)
        formulas_combined.extend(formulas)

    df_values = pd.DataFrame(data_combined)
    df_formulas = pd.DataFrame(formulas_combined)

    return df_values, df_formulas

# Path to the Excel file
file_path = 'large_file_with_formulas.xlsx'

# Extract values and formulas using multiprocessing
df_values, df_formulas = extract_formulas_multiprocessing(file_path)

# Display the DataFrames
print("Values:")
print(df_values.head())

print("\nFormulas:")
print(df_formulas.head())
